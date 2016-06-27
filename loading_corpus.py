import nltk
#from nltk.book import *
import openpyxl
from openpyxl import Workbook

'''
Loading Txt file data

f=open('/home/divyat/Desktop/RTE/Data/SemEval2016-Task6-testdata/SemEval2016-Task6-subtaskA-testdata.txt','rU')
text=f.read()
text1=text.split()
abstracts=nltk.Text(text1)
'''

#Loading xlsx Sheets to create workbook object:collection of worksheet objects
wb=openpyxl.load_workbook('/home/divyat/Desktop/RTE/Data/TrainingData/training_data_1.xlsx')
name=wb.get_sheet_names()
print name

#Load data into sheet to create a worksheet object
sheet=wb.get_sheet_by_name('training_data_1.xlsx')

print 'Type',
print str(type(sheet))

print 'Title',
print str(sheet.title)

#Accessing values from cells
cell=sheet['B3']
print cell.value

cell=sheet.cell(row=1,column=2)
print cell.value

for i in range(2,10):
	cell=sheet.cell(row=i,column=2)
	print cell.value

